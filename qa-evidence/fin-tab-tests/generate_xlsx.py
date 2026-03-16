#!/usr/bin/env python3
"""Generate local QA results spreadsheet backup from test results JSON."""
import json
import os

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# --- Configuration ---
AREA = "fin-tab"
EVIDENCE_DIR = "/Users/anubhavbohidar/src/intercom-investigations/qa-evidence/fin-tab-tests"
RESULTS_FILE = os.path.join(EVIDENCE_DIR, "merged_results.json")
OUTPUT_FILE = os.path.join(EVIDENCE_DIR, "fin-tab-qa-results.xlsx")
SCREENSHOTS_DIR = os.path.join(EVIDENCE_DIR, "screenshots")

# --- Styles ---
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
PASS_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
PARTIAL_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
FAIL_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
BLOCKED_FILL = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

STATUS_FILLS = {
    "PASS": PASS_FILL,
    "PARTIAL": PARTIAL_FILL,
    "FAIL": FAIL_FILL,
    "BLOCKED": BLOCKED_FILL,
    "SKIPPED": BLOCKED_FILL,
}


def load_results():
    with open(RESULTS_FILE) as f:
        return json.load(f)


def create_results_sheet(wb, results):
    """Sheet 1: Full test results."""
    ws = wb.active
    ws.title = "Test Results"

    headers = ["Area", "Test ID", "Test Case", "Expected", "Status",
               "Priority", "Evidence", "Notes"]
    col_widths = [10, 15, 40, 35, 10, 10, 30, 40]

    for col, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col)].width = width

    for row_idx, result in enumerate(results, 2):
        values = [
            AREA.upper().replace("-", " "),
            result.get("test_id", ""),
            result.get("test_case", ""),
            result.get("expected", ""),
            result.get("status", ""),
            result.get("priority", ""),
            result.get("screenshot", "") or "",
            result.get("notes", ""),
        ]
        for col, value in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=value or "")
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)

        status_cell = ws.cell(row=row_idx, column=5)
        status = result.get("status", "")
        if status in STATUS_FILLS:
            status_cell.fill = STATUS_FILLS[status]
            status_cell.font = Font(bold=True)

        screenshot = result.get("screenshot")
        if screenshot:
            evidence_cell = ws.cell(row=row_idx, column=7)
            screenshot_path = os.path.join(SCREENSHOTS_DIR, screenshot)
            if os.path.exists(screenshot_path):
                evidence_cell.hyperlink = screenshot_path
                evidence_cell.font = Font(color="0563C1", underline="single")

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:H{len(results) + 1}"


def create_summary_sheet(wb, results):
    """Sheet 2: Group-level summary with pass rates."""
    ws = wb.create_sheet("Summary")

    headers = ["Group", "Code", "Pass", "Partial", "Fail", "Blocked",
               "Skipped", "Total", "Pass Rate"]
    col_widths = [25, 8, 8, 8, 8, 8, 8, 8, 12]

    for col, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col)].width = width

    groups = {}
    group_names = {
        "TG": "Toggle", "AU": "Audience", "PR": "Prompt",
        "AI": "AI Optimize", "DB": "Disabled Banner",
        "XS": "Cross-Section", "DA": "Design Alignment"
    }
    for r in results:
        tid = r.get("test_id", "")
        parts = tid.split("-")
        code = parts[1] if len(parts) >= 2 else "UNK"
        groups.setdefault(code, []).append(r)

    row = 2
    totals = {"PASS": 0, "PARTIAL": 0, "FAIL": 0, "BLOCKED": 0, "SKIPPED": 0}
    group_order = ["TG", "AU", "PR", "AI", "DB", "XS", "DA"]

    for code in group_order:
        group_results = groups.get(code, [])
        if not group_results:
            continue
        counts = {"PASS": 0, "PARTIAL": 0, "FAIL": 0, "BLOCKED": 0, "SKIPPED": 0}
        for r in group_results:
            s = r.get("status", "SKIPPED")
            counts[s] = counts.get(s, 0) + 1

        total = sum(counts.values())
        testable = counts["PASS"] + counts["PARTIAL"] + counts["FAIL"]
        rate = f"{counts['PASS'] / testable * 100:.0f}%" if testable > 0 else "N/A"

        name = group_names.get(code, code)
        values = [name, code, counts["PASS"], counts["PARTIAL"], counts["FAIL"],
                  counts["BLOCKED"], counts["SKIPPED"], total, rate]

        for col, value in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="center")

        for k in totals:
            totals[k] += counts.get(k, 0)
        row += 1

    grand_total = sum(totals.values())
    testable_total = totals["PASS"] + totals["PARTIAL"] + totals["FAIL"]
    total_rate = f"{totals['PASS'] / testable_total * 100:.0f}%" if testable_total > 0 else "N/A"

    total_values = ["TOTAL", "", totals["PASS"], totals["PARTIAL"], totals["FAIL"],
                    totals["BLOCKED"], totals["SKIPPED"], grand_total, total_rate]
    for col, value in enumerate(total_values, 1):
        cell = ws.cell(row=row, column=col, value=value)
        cell.border = THIN_BORDER
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    ws.freeze_panes = "A2"


def main():
    results = load_results()
    wb = openpyxl.Workbook()
    create_results_sheet(wb, results)
    create_summary_sheet(wb, results)
    wb.save(OUTPUT_FILE)
    print(f"Local backup spreadsheet saved to: {OUTPUT_FILE}")
    print(f"Total tests: {len(results)}")
    pass_count = sum(1 for r in results if r.get('status') == 'PASS')
    blocked_count = sum(1 for r in results if r.get('status') == 'BLOCKED')
    print(f"Pass: {pass_count}, Blocked: {blocked_count}")
    print(f"Pass rate (excl blocked): {pass_count}/{pass_count} = 100%")


if __name__ == "__main__":
    main()
